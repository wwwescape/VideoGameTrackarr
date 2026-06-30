"""Loads the curated hardware reference dataset from data/hardware/*.xlsx (one master
hardware database per brand — Nintendo, Sony, Xbox) into the hardware_reference_entries
table. Powers the Brand/Console/Variant cascades on the Add Device/Add Accessory
(Predefined) forms and the "rich" reference data shown on the Device/Accessory detail
pages.

Usage:
    python -m scripts.import_hardware_reference
    python -m scripts.import_hardware_reference --path "../data/hardware"

Safe to re-run: rows are upserted by `official_name` (globally unique across every source
file) — re-running after the source spreadsheets are updated just updates existing rows
instead of duplicating them.
"""

import argparse
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.core.config import REPO_ROOT
from app.db.session import session_scope
from app.repositories import hardware_reference_repository

DEFAULT_DIR = REPO_ROOT / "data" / "hardware"


def _data_rows(ws: Worksheet):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            yield row


def _as_year_text(value: object) -> str | None:
    """Reduces a Release Date cell to just its 4-digit year — the source spreadsheets mix
    real Excel date cells with bare-year text/numbers, but the app only ever needs the year."""
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return str(value.year)
    text = str(value).strip()
    if not text:
        return None
    return text[:4] if text[:4].isdigit() else text


def import_from_workbook(session: Session, path: Path) -> int:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = workbook[workbook.sheetnames[0]]

    count = 0
    for row in _data_rows(ws):
        (
            brand,
            family,
            generation,
            generation_short,
            artefact,
            official_name,
            category,
            entry_type,
            release_date,
            discontinued,
            compatibility,
            summary,
        ) = row

        hardware_reference_repository.upsert(
            session,
            official_name=official_name,
            brand=brand,
            family=family or None,
            generation=generation,
            generation_short=generation_short or None,
            artefact=artefact,
            category=category,
            type=entry_type,
            release_date=_as_year_text(release_date),
            discontinued=str(discontinued).strip().lower() == "yes",
            compatibility=compatibility or None,
            summary=summary or None,
        )
        count += 1

    workbook.close()
    print(f"{path.name}: upserted {count} row(s)")
    return count


def import_from_directory(session: Session, directory: Path) -> int:
    total = 0
    for path in sorted(directory.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        total += import_from_workbook(session, path)
    print(f"Hardware reference entries: upserted {total} row(s) total")
    return total


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--path", default=str(DEFAULT_DIR), help="Directory containing the source .xlsx files")
    args = parser.parse_args()

    with session_scope() as session:
        import_from_directory(session, Path(args.path))


if __name__ == "__main__":
    main()
